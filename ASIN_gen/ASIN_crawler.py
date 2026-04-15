import os
import asyncio
import re
import io
import pandas as pd
from datetime import datetime
from playwright.async_api import async_playwright, Route
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

# ================= 🛡️ 0. 环境兼容性修复 =================
if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

# 获取脚本所在的绝对路径
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ================= 🎨 1. 商业级调色盘 =================
COLOR_MAP = {
    "HEADER": "2C3E50", 
    "NORMAL": "FFFFFF",          # 正常
    "SOFT_OOS": "E2EFDA",        # 预售
    "HARD_OOS": "FFE699",        # 断货
    "DOG": "F8CBAD",             # 变狗
    "REDIRECT_NORMAL": "D9E1F2",  # 跳转正常
    "REDIRECT_SOFT_OOS": "D0ECE7",# 跳转预售
    "REDIRECT_HARD_OOS": "E4DFEC",# 跳转断货
    "ERROR": "F2D7D5"             # 异常
}
STYLES = {k: PatternFill("solid", fgColor=v) for k, v in COLOR_MAP.items()}
SOFT_BORDER = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), 
                      top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)

# ================= 🚀 2. 激进拦截与伪装 =================
DISCARD_ASSETS = ["image", "media", "font", "texttrack", "object", "beacon", "csp_report"]
BLOCK_URLS = ["amazon-adsystem", "google-analytics", "doubleclick", "forensics", "metrics", "log"]

async def block_resources(route: Route):
    url = route.request.url.lower()
    if route.request.resource_type in DISCARD_ASSETS or any(d in url for d in BLOCK_URLS):
        await route.abort()
    else:
        await route.continue_()

# ================= ⚙️ 3. 核心爬虫引擎 =================
async def scrape_single_url(url: str, context, semaphore: asyncio.Semaphore):
    async with semaphore:
        # 提取原始输入 ASIN
        origin_asin = (re.search(r'/(?:dp|gp/product)/([A-Z0-9]{10})', url, re.IGNORECASE) or [None, "UNKNOWN"])[1].upper()
        
        result = {
            "原始ASIN": origin_asin, "实际ASIN": "/", "最终状态": "等待中", "状态KEY": "ERROR",
            "品牌": "/", "价格": "/", "上月的销量": "/", "星级": "/", "评价数": "/", "标题": "/",
            **{f"F{i}": "/" for i in range(1, 11)}, "URL": url
        }

        page = await context.new_page()
        try:
            # 抹除浏览器特征
            await page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            await page.route("**/*", block_resources)
            
            response = await page.goto(url, timeout=60000, wait_until="domcontentloaded")
            
            # 竞速加载探测
            try:
                await page.wait_for_selector("#productTitle, #ASIN, .error_handler", timeout=5000)
            except: pass

            # 1. 基础异常判定
            if response.status == 404 or "Robot Check" in await page.title():
                result.update({"最终状态": "变狗失效", "状态KEY": "DOG"})
                return result

            # 2. 跳转判定
            real_asin = await page.get_attribute("#ASIN", "value")
            if not real_asin:
                real_asin = (re.search(r'/dp/([A-Z0-9]{10})', page.url, re.IGNORECASE) or [None, origin_asin])[1]
            
            real_asin = real_asin.upper()
            result["实际ASIN"] = real_asin
            is_redirect = (real_asin != origin_asin)

            # 3. 状态解析
            body_text = (await page.inner_text("body")).lower()
            has_buy_btn = await page.locator("#add-to-cart-button, #buy-now-button").count() > 0

            if "currently unavailable" in body_text:
                result["最终状态"], result["状态KEY"] = ("跳转_断货" if is_redirect else "彻底断货"), ("REDIRECT_HARD_OOS" if is_redirect else "HARD_OOS")
            elif "temporarily out of stock" in body_text:
                status_text = "跳转_预售" if is_redirect else "预售可订"
                status_key = "REDIRECT_SOFT_OOS" if is_redirect else "SOFT_OOS"
                if not has_buy_btn: # 没按钮的预售算断货
                    status_text, status_key = ("跳转_断货" if is_redirect else "彻底断货"), ("REDIRECT_HARD_OOS" if is_redirect else "HARD_OOS")
                result["最终状态"], result["状态KEY"] = status_text, status_key
            else:
                result["最终状态"], result["状态KEY"] = ("跳转_正常" if is_redirect else "正常在线"), ("REDIRECT_NORMAL" if is_redirect else "NORMAL")

            # 4. 字段解析
            title_el = page.locator("#productTitle").first
            if await title_el.count() > 0: result["标题"] = (await title_el.inner_text()).strip()

            brand_el = page.locator("#bylineInfo").first
            if await brand_el.count() > 0:
                result["品牌"] = (await brand_el.inner_text()).replace("Visit the ", "").replace(" Store", "").strip()

            p_whole = page.locator(".a-price-whole").first
            if await p_whole.count() > 0:
                w = (await p_whole.inner_text()).strip().replace('.', '').replace('\n', '')
                f = (await page.locator(".a-price-fraction").first.inner_text()) if await page.locator(".a-price-fraction").count() > 0 else "00"
                result["价格"] = f"{w}.{f}"

            sales_el = page.locator("#social-proofing-faceout-title-tk_bought").first
            if await sales_el.count() > 0:
                m = re.search(r'(\d+[\d,Kk\+]*\+?)', await sales_el.inner_text())
                if m: result["上月销量"] = m.group(1)

            features = await page.locator("#feature-bullets ul li span.a-list-item").all_inner_texts()
            clean_f = [f.strip() for f in features if f.strip() and "Make sure this fits" not in f]
            for i in range(min(10, len(clean_f))): result[f"F{i+1}"] = clean_f[i]

            print(f"✅ {origin_asin} -> {result['最终状态']}")
        except Exception as e:
            result["最终状态"], result["状态KEY"] = "抓取异常", "ERROR"
            print(f"❌ {origin_asin} 异常: {str(e)[:30]}")
        finally:
            await page.close()
            return result

# ================= 📊 4. 稳健的 Excel 保存逻辑 =================
def save_report(results, site):
    ts = datetime.now().strftime("%m%d_%H%M")
    # 使用绝对路径生成报告
    filename = os.path.join(BASE_DIR, f"Amazon_Monitor_{site}.xlsx")
    
    wb = openpyxl.load_workbook(filename) if os.path.exists(filename) else openpyxl.Workbook()
    ws = wb.active if "Sheet" in wb.sheetnames else wb.create_sheet(title=f"Intel_{ts}")
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1: del wb["Sheet"]

    # 1. 统计看板
    keys = ["NORMAL", "SOFT_OOS", "HARD_OOS", "DOG", "REDIRECT_NORMAL", "REDIRECT_SOFT_OOS", "REDIRECT_HARD_OOS", "ERROR"]
    labels = ["📊 总数", "正常", "预售", "断货", "变狗", "跳_正", "跳_预", "跳_断", "异常"]
    
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 30

    for i, label in enumerate(labels, 1):
        cell = ws.cell(2, i, label)
        cell.fill = STYLES["HEADER"]; cell.font = Font(color="FFFFFF", bold=True); cell.alignment = CENTER_ALIGN
    
    # 计算统计值
    counts = {k: sum(1 for r in results if r["状态KEY"] == k) for k in keys}
    ws.cell(3, 1, len(results)).font = Font(bold=True, size=12); ws.cell(3, 1).alignment = CENTER_ALIGN
    for i, k in enumerate(keys, 2):
        cell = ws.cell(3, i, counts[k])
        cell.fill = STYLES.get(k, STYLES["NORMAL"]); cell.alignment = CENTER_ALIGN; cell.font = Font(bold=True)

    # 2. 数据表头
    headers = ["原始ASIN", "实际ASIN", "最终状态", "品牌", "价格", "上月销量", "标题"] + [f"F{i}" for i in range(1, 11)] + ["URL"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(5, i, h)
        cell.fill = STYLES["HEADER"]; cell.font = Font(color="FFFFFF", bold=True); cell.alignment = CENTER_ALIGN

    # 3. 写入数据
    for row_idx, res in enumerate(results, 6):
        fill = STYLES.get(res["状态KEY"], STYLES["NORMAL"])
        for col_idx, h in enumerate(headers, 1):
            val = res.get(h, "/")
            cell = ws.cell(row_idx, col_idx, val)
            cell.fill = fill; cell.border = SOFT_BORDER; cell.alignment = LEFT_ALIGN if col_idx > 6 else CENTER_ALIGN

    # 4. 自动列宽
    widths = [15, 15, 12, 15, 10, 12, 40] + [30]*10 + [20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(filename)
    print(f"\n📊 报告已生成: {filename}")

# ================= 🚀 5. 主程序 =================
async def main():
    # 动态获取 urls.txt 的绝对路径
    urls_file = os.path.join(BASE_DIR, "urls.txt")
    
    if not os.path.exists(urls_file):
        print(f"❌ 错误：找不到文件 {urls_file}")
        print(f"请检查文件是否在：{BASE_DIR}")
        return
    
    with open(urls_file, "r", encoding="utf-8") as f:
        urls = [l.strip() for l in f if l.strip()]
    
    if not urls: print("⚠️ urls.txt 是空的"); return

    site = input("👉 请选择站点 (us/uk/de/jp...): ").strip().lower()
    print(f"\n⚡️ 启动 8 并发引擎 | 待处理: {len(urls)} 条")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        
        # 动态获取 auth.json 的绝对路径
        auth_file = os.path.join(BASE_DIR, f"auth_{site}.json")
        
        # 伪装环境
        context = await browser.new_context(
            storage_state=auth_file if os.path.exists(auth_file) else None,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
            locale="en-US", timezone_id="America/New_York"
        )
        
        sem = asyncio.Semaphore(8)
        tasks = [scrape_single_url(u, context, sem) for u in urls]
        results = await asyncio.gather(*tasks)
        
        await browser.close()
        save_report(results, site)

if __name__ == "__main__":
    asyncio.run(main())